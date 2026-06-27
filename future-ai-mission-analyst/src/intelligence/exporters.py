"""Report exporters — Excel (.xlsx) and PDF.

The mission layer already emits Markdown + JSON. These helpers add the two
formats operators actually circulate: a spreadsheet of the event timeline and a
printable PDF mission summary.

Every function returns ``bytes`` so the Streamlit ``download_button`` can serve
them straight to the browser, and :class:`MissionIntelligence.export` can also
write them to disk. Heavy deps (openpyxl, reportlab) are imported lazily so the
rest of the app loads even if they are absent.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Dict, List, Optional


def _now_str() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")


# ──────────────────────────────────────────────────────── generic text PDF
def sections_to_pdf(title: str, sections: List["tuple[str, str]"]) -> bytes:
    """Render a simple titled PDF from ``(heading, body)`` sections.

    Used for the single/dual-image EO/IR analysis reports.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=18 * mm,
                            rightMargin=18 * mm, topMargin=16 * mm,
                            bottomMargin=16 * mm, title=title)
    styles = getSampleStyleSheet()
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=10, leading=14)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8,
                           textColor=colors.grey)

    def esc(s: str) -> str:
        return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;"))

    story: list = [Paragraph(esc(title), styles["Title"]),
                   Paragraph(f"Generated: {_now_str()}", small), Spacer(1, 6)]
    for heading, text in sections:
        if not text:
            continue
        story.append(Paragraph(esc(heading), h2))
        story.append(Paragraph(esc(text), body))
    doc.build(story)
    return buf.getvalue()


# ───────────────────────────────────────────────────────────────────── Excel
def timeline_to_xlsx(
    timeline: List[Dict],
    stats: Optional[Dict] = None,
    final_summary: str = "",
    title: str = "Mission Event Timeline",
) -> bytes:
    """Render the event timeline (+ a stats sheet) to an .xlsx workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # -- Sheet 1: Timeline --------------------------------------------------
    ws = wb.active
    ws.title = "Timeline"
    header = ["#", "Time", "Type", "Kind", "Detail"]
    ws.append(header)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(header, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(horizontal="center")

    for i, row in enumerate(timeline, start=1):
        ws.append([
            i,
            row.get("time", ""),
            row.get("type", ""),
            row.get("kind", ""),
            row.get("detail", ""),
        ])

    widths = [5, 12, 20, 10, 90]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.cell(row=1, column=1)  # ensure cells exist
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    # -- Sheet 2: Summary / Stats ------------------------------------------
    ss = wb.create_sheet("Summary")
    ss.append([title])
    ss["A1"].font = Font(bold=True, size=14)
    ss.append([f"Generated: {_now_str()}"])
    ss.append([])
    if final_summary:
        ss.append(["Executive Summary"])
        ss[f"A{ss.max_row}"].font = Font(bold=True)
        ss.append([final_summary])
        ss[f"A{ss.max_row}"].alignment = Alignment(wrap_text=True, vertical="top")
        ss.append([])
    if stats:
        ss.append(["Statistic", "Value"])
        hr = ss.max_row
        for col in (1, 2):
            ss.cell(row=hr, column=col).font = Font(bold=True)
        ss.append(["Unique tracked objects", stats.get("unique_objects", 0)])
        ss.append(["Timeline entries", stats.get("timeline_len", 0)])
        ss.append([])
        by_class = stats.get("by_class", {}) or {}
        if by_class:
            ss.append(["Objects by class", "Count"])
            for col in (1, 2):
                ss.cell(row=ss.max_row, column=col).font = Font(bold=True)
            for cls, n in by_class.items():
                ss.append([cls, n])
            ss.append([])
        events = stats.get("events", {}) or {}
        if events:
            ss.append(["Event type", "Count"])
            for col in (1, 2):
                ss.cell(row=ss.max_row, column=col).font = Font(bold=True)
            for ev, n in events.items():
                ss.append([ev, n])
    ss.column_dimensions["A"].width = 28
    ss.column_dimensions["B"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────── PDF
def mission_to_pdf(
    timeline: List[Dict],
    stats: Optional[Dict] = None,
    final_summary: str = "",
    title: str = "Mission Intelligence Report",
    intel_feed: Optional[List[str]] = None,
    snapshots: Optional[List[Dict]] = None,
) -> bytes:
    """Render a printable PDF mission report.

    ``snapshots`` (optional): list of ``{"jpeg": bytes, "identity": int,
    "class_name": str, "confidence": float, "time": str}`` — rendered as a
    "Recorded Objects" thumbnail gallery.
    """
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=16 * mm, bottomMargin=16 * mm,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=10, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], alignment=TA_LEFT,
                          fontSize=9.5, leading=13)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8,
                           textColor=colors.grey)

    def esc(s: str) -> str:
        return (str(s).replace("&", "&amp;").replace("<", "&lt;")
                .replace(">", "&gt;"))

    story: list = [Paragraph(esc(title), h1),
                   Paragraph(f"Generated: {_now_str()}", small),
                   Spacer(1, 6)]

    story.append(Paragraph("Executive Summary", h2))
    story.append(Paragraph(esc(final_summary or "No VLM summary generated."), body))

    if stats:
        story.append(Paragraph("Statistics", h2))
        rows = [["Metric", "Value"],
                ["Unique tracked objects", str(stats.get("unique_objects", 0))],
                ["Timeline entries", str(stats.get("timeline_len", 0))]]
        for cls, n in (stats.get("by_class", {}) or {}).items():
            rows.append([f"Class · {cls}", str(n)])
        for ev, n in (stats.get("events", {}) or {}).items():
            rows.append([f"Event · {ev}", str(n)])
        t = Table(rows, colWidths=[90 * mm, 80 * mm], hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
        ]))
        story.append(t)

    # -- Recorded objects: thumbnail gallery -------------------------------
    if snapshots:
        story.append(Paragraph("Recorded Objects", h2))
        cap = ParagraphStyle("cap", parent=small, alignment=TA_CENTER,
                             fontSize=7, leading=8, textColor=colors.black)
        cols = 5
        thumb_w = 30 * mm
        cells: list = []
        for s in snapshots[:60]:
            try:
                img = Image(io.BytesIO(s["jpeg"]))
            except Exception:  # noqa: BLE001 — skip an unreadable thumbnail
                continue
            aspect = (img.imageHeight / img.imageWidth) if img.imageWidth else 1.0
            img.drawWidth = thumb_w
            img.drawHeight = max(8 * mm, min(42 * mm, thumb_w * aspect))
            caption = (f"#{s.get('identity')} {esc(s.get('class_name', ''))}<br/>"
                       f"{esc(s.get('time', ''))} · {s.get('confidence', 0):.2f}")
            cells.append([img, Paragraph(caption, cap)])
        rows = [cells[i:i + cols] for i in range(0, len(cells), cols)]
        for r in rows:
            while len(r) < cols:
                r.append("")
        if rows:
            gt = Table(rows, colWidths=[(170 / cols) * mm] * cols, hAlign="LEFT")
            gt.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(gt)

    story.append(Paragraph("Event Timeline", h2))
    if timeline:
        tl_rows = [["Time", "Type", "Detail"]]
        for row in timeline:
            tl_rows.append([
                Paragraph(esc(row.get("time", "")), body),
                Paragraph(esc(row.get("type", "")), body),
                Paragraph(esc(row.get("detail", "")), body),
            ])
        t = Table(tl_rows, colWidths=[22 * mm, 38 * mm, 114 * mm],
                  repeatRows=1, hAlign="LEFT")
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
        ]))
        story.append(t)
    else:
        story.append(Paragraph("No events recorded.", body))

    if intel_feed:
        story.append(Paragraph("Intel Feed", h2))
        for msg in intel_feed:
            story.append(Paragraph("• " + esc(msg), body))

    doc.build(story)
    return buf.getvalue()
